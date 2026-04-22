from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

DAYS = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
DAY_INDEX = {d: i for i, d in enumerate(DAYS)}
MAX_SLOTS = 13
MINUTES_PER_DAY = 1440


@dataclass
class DaySchedule:
    slots: list[tuple[int, float]] = field(default_factory=list)


@dataclass
class RoomSchedule:
    room: str
    days: dict[str, DaySchedule] = field(default_factory=dict)


def _parse_time(value) -> int:
    import datetime
    if isinstance(value, str):
        h, m = value.strip().split(":")
        return int(h) * 60 + int(m)
    if isinstance(value, float):
        total = round(value * MINUTES_PER_DAY)
        return total % MINUTES_PER_DAY or MINUTES_PER_DAY
    if isinstance(value, datetime.datetime):
        return value.hour * 60 + value.minute
    if isinstance(value, datetime.time):
        return value.hour * 60 + value.minute
    raise ValueError(f"Cannot parse time: {value!r}")


def _parse_group_label(label: str) -> list[str]:
    """Expand a group label like 'Mo-Fr', 'Sa+So', 'Mi' to a list of day tokens."""
    s = label.strip()
    if not s:
        raise ValueError("empty group label")
    parts: list[str] = []
    for token in s.replace(" ", "").split("+"):
        if "-" in token:
            a, b = token.split("-", 1)
            if a not in DAY_INDEX or b not in DAY_INDEX:
                raise ValueError(f"unknown day in range '{token}'")
            i, j = DAY_INDEX[a], DAY_INDEX[b]
            if i > j:
                raise ValueError(f"inverted range '{token}'")
            parts.extend(DAYS[i:j + 1])
        else:
            if token not in DAY_INDEX:
                raise ValueError(f"unknown day '{token}'")
            parts.append(token)
    return parts


def parse_xlsx(path: Path) -> list[RoomSchedule]:
    wb = openpyxl.load_workbook(path, data_only=True)
    rooms = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rooms.append(_parse_sheet(sheet_name, ws))
    return rooms


def _parse_sheet(room: str, ws) -> RoomSchedule:
    rows = list(ws.iter_rows(values_only=True))

    # Collect blocks: each block = group label row + header row + data rows.
    # A block starts when column A holds a group label (not 'von'); the next row
    # must be the header ('von','bis','Temp'); data rows follow until an empty
    # row or the next group label.
    day_slots: dict[str, list[tuple[int, float]]] = {d: [] for d in DAYS}
    assigned_days: set[str] = set()

    i = 0
    n = len(rows)
    while i < n:
        first = rows[i][0] if rows[i] else None
        if first is None or (isinstance(first, str) and not first.strip()):
            i += 1
            continue
        label = str(first).strip()
        if label.lower() == "von":
            raise ValueError(
                f"Sheet '{room}' row {i+1}: header 'von' without preceding group label"
            )

        # Group label row — next row must be header
        if i + 1 >= n or not rows[i + 1] or str(rows[i + 1][0] or "").strip().lower() != "von":
            raise ValueError(
                f"Sheet '{room}' row {i+1}: group '{label}' is not followed by a 'von/bis/Temp' header"
            )

        try:
            group_days = _parse_group_label(label)
        except ValueError as e:
            raise ValueError(f"Sheet '{room}' row {i+1}: {e}") from None

        dup = [d for d in group_days if d in assigned_days]
        if dup:
            raise ValueError(
                f"Sheet '{room}' row {i+1}: group '{label}' redefines days already set: {dup}"
            )

        # Read data rows
        j = i + 2
        block_slots: list[tuple[int, float]] = []
        while j < n:
            row = rows[j]
            v = row[0] if row else None
            if v is None or (isinstance(v, str) and not v.strip()):
                break
            # A new group label would be a non-time string in column A
            if isinstance(v, str):
                # If it looks like HH:MM, treat as time; otherwise stop block
                try:
                    _parse_time(v)
                except ValueError:
                    break
            bis = row[1] if len(row) > 1 else None
            tmp = row[2] if len(row) > 2 else None
            if bis is None or tmp is None:
                raise ValueError(
                    f"Sheet '{room}' row {j+1}: missing 'bis' or 'Temp' value"
                )
            try:
                endtime = _parse_time(bis)
            except ValueError as e:
                raise ValueError(f"Sheet '{room}' row {j+1}: bis: {e}") from None
            if endtime == 0:
                endtime = MINUTES_PER_DAY
            try:
                temperature = float(tmp)
            except (TypeError, ValueError):
                raise ValueError(
                    f"Sheet '{room}' row {j+1}: cannot parse temperature {tmp!r}"
                ) from None
            block_slots.append((endtime, temperature))
            j += 1

        if not block_slots:
            raise ValueError(f"Sheet '{room}' group '{label}': no data rows")
        block_slots.sort(key=lambda x: x[0])
        if block_slots[-1][0] != MINUTES_PER_DAY:
            raise ValueError(
                f"Sheet '{room}' group '{label}': last endtime must be 24:00, got {block_slots[-1][0]}"
            )
        if len(block_slots) > MAX_SLOTS:
            raise ValueError(
                f"Sheet '{room}' group '{label}': {len(block_slots)} slots exceed max {MAX_SLOTS}"
            )

        for day in group_days:
            day_slots[day] = list(block_slots)
            assigned_days.add(day)

        i = j

    missing = [d for d in DAYS if d not in assigned_days]
    if missing:
        raise ValueError(f"Sheet '{room}': days not covered by any group: {missing}")

    rs = RoomSchedule(room=room)
    for day in DAYS:
        rs.days[day] = DaySchedule(slots=day_slots[day])
    return rs
